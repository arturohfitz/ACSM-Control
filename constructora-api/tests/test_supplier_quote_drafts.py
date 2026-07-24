import unittest
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.supplier_quote_drafts import build_quote_template, parse_quote_template


def _quote_link(*, supplier_id: int = 27, rfq_number: str = "SC-TEST-001"):
    items = [
        SimpleNamespace(
            id=101,
            source_code="MAT-001",
            description="Cemento gris",
            quantity=12.5,
            unit="SACO",
        ),
        SimpleNamespace(
            id=102,
            source_code="MAT-002",
            description="Varilla 3/8",
            quantity=48,
            unit="PZA",
        ),
    ]
    return SimpleNamespace(
        supplier_id=supplier_id,
        supplier=SimpleNamespace(name="Proveedor de prueba"),
        rfq=SimpleNamespace(rfq_number=rfq_number, items=items),
    )


class SupplierQuoteDraftTemplateTest(unittest.TestCase):
    def test_template_round_trip_preserves_supplier_rfq_and_item_ids(self) -> None:
        link = _quote_link()
        workbook = load_workbook(BytesIO(build_quote_template(link)))
        sheet = workbook["Cotizacion"]
        sheet["B4"] = "COT-2026-001"
        sheet["B5"] = "2026-08-15"
        sheet["B7"] = 5
        sheet["B9"] = 100
        sheet["B10"] = 250
        sheet["B11"] = 480
        sheet["F14"] = 120.50
        sheet["G14"] = 3
        sheet["H14"] = "Entrega en obra"
        sheet["F15"] = 35.75

        output = BytesIO()
        workbook.save(output)
        parsed = parse_quote_template(output.getvalue(), link)

        self.assertEqual(parsed.quote_number, "COT-2026-001")
        self.assertEqual(parsed.delivery_days, 5)
        self.assertEqual(parsed.discount, 100)
        self.assertEqual(parsed.shipping_cost, 250)
        self.assertEqual(parsed.tax_amount, 480)
        self.assertEqual([item.rfq_item_id for item in parsed.items], [101, 102])
        self.assertEqual(parsed.items[0].unit_price, 120.5)
        self.assertEqual(parsed.items[0].delivery_days, 3)
        self.assertEqual(parsed.items[0].notes, "Entrega en obra")

    def test_template_cannot_be_submitted_by_another_supplier(self) -> None:
        original_link = _quote_link(supplier_id=27)
        foreign_link = _quote_link(supplier_id=99)
        workbook = load_workbook(BytesIO(build_quote_template(original_link)))
        workbook["Cotizacion"]["B4"] = "COT-FOREIGN"
        workbook["Cotizacion"]["F14"] = 100
        output = BytesIO()
        workbook.save(output)

        with self.assertRaisesRegex(ValueError, "otro proveedor"):
            parse_quote_template(output.getvalue(), foreign_link)

    def test_template_cannot_be_submitted_for_another_rfq(self) -> None:
        original_link = _quote_link(rfq_number="SC-TEST-001")
        foreign_link = _quote_link(rfq_number="SC-TEST-999")
        workbook = load_workbook(BytesIO(build_quote_template(original_link)))
        workbook["Cotizacion"]["B4"] = "COT-FOREIGN"
        workbook["Cotizacion"]["F14"] = 100
        output = BytesIO()
        workbook.save(output)

        with self.assertRaisesRegex(ValueError, "otra solicitud"):
            parse_quote_template(output.getvalue(), foreign_link)


if __name__ == "__main__":
    unittest.main()
