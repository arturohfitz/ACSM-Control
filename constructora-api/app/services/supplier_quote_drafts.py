from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models import (
    SupplierQuoteDraft,
    SupplierQuoteDraftItem,
    SupplierQuoteUpload,
    SupplierRFQSupplier,
)
from app.schemas.purchasing import SupplierQuoteDraftInput


TEMPLATE_MARKER = "ACSM_QUOTE_TEMPLATE"
TEMPLATE_VERSION = "1"
ITEM_HEADER_ROW = 13


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"))


def _decimal(value: object, *, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"Valor numerico invalido: {value}") from exc


def _integer(value: object, *, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Numero entero invalido: {value}") from exc
    if parsed < 0:
        raise ValueError(f"El valor no puede ser negativo: {value}")
    return parsed


def create_quote_draft(
    db: Session,
    *,
    link: SupplierRFQSupplier,
    upload: SupplierQuoteUpload,
    payload: SupplierQuoteDraftInput,
    source_type: str,
    confidence: Decimal,
    parser_version: str | None = None,
    initial_errors: list[str] | None = None,
) -> SupplierQuoteDraft:
    rfq_items = {item.id: item for item in link.rfq.items}
    submitted_items = {item.rfq_item_id: item for item in payload.items}
    if len(submitted_items) != len(payload.items):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cotizacion contiene partidas repetidas",
        )
    unknown_ids = sorted(set(submitted_items) - set(rfq_items))
    if unknown_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cotizacion contiene partidas ajenas a la solicitud",
        )

    errors = list(initial_errors or [])
    missing_count = len(rfq_items) - len(submitted_items)
    if missing_count:
        errors.append(f"Faltan precios para {missing_count} partida(s)")

    subtotal = Decimal("0")
    draft = SupplierQuoteDraft(
        company_id=link.rfq.company_id,
        rfq_id=link.rfq_id,
        rfq_supplier_id=link.id,
        supplier_id=link.supplier_id,
        upload_id=upload.id,
        status="review_required",
        source_type=source_type,
        parser_version=parser_version,
        confidence=confidence,
        quote_number=payload.quote_number.strip(),
        received_at=date.today(),
        valid_until=payload.valid_until,
        currency=payload.currency.upper(),
        delivery_days=payload.delivery_days,
        payment_terms_days=payload.payment_terms_days,
        discount=_money(payload.discount),
        shipping_cost=_money(payload.shipping_cost),
        tax_amount=_money(payload.tax_amount),
        notes=payload.notes,
        validation_errors=errors,
    )
    db.add(draft)
    db.flush()

    for rfq_item in link.rfq.items:
        submitted = submitted_items.get(rfq_item.id)
        unit_price = submitted.unit_price if submitted is not None else None
        line_total = _money(rfq_item.quantity * unit_price) if unit_price is not None else Decimal("0")
        subtotal += line_total
        db.add(
            SupplierQuoteDraftItem(
                draft_id=draft.id,
                rfq_item_id=rfq_item.id,
                material_id=rfq_item.material_id,
                description=rfq_item.description,
                unit=rfq_item.unit,
                quantity=rfq_item.quantity,
                unit_price=unit_price,
                line_total=line_total,
                delivery_days=submitted.delivery_days if submitted else None,
                notes=submitted.notes if submitted else None,
                confidence=confidence if submitted else Decimal("0"),
                match_method="rfq_item_id" if submitted else "missing",
            )
        )

    draft.subtotal = _money(subtotal)
    draft.total = _money(
        max(
            Decimal("0"),
            draft.subtotal - draft.discount + draft.shipping_cost + draft.tax_amount,
        )
    )
    upload.status = "review_required"
    db.flush()
    return draft


def build_quote_template(link: SupplierRFQSupplier) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cotizacion"
    metadata = workbook.create_sheet("_ACSM")
    metadata.sheet_state = "hidden"
    metadata.append([TEMPLATE_MARKER, TEMPLATE_VERSION])
    metadata.append(["RFQ", link.rfq.rfq_number])
    metadata.append(["SUPPLIER_ID", link.supplier_id])

    sheet["A1"] = "Cotizacion para ACSM Control"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B5F94")
    sheet.merge_cells("A1:H1")
    sheet["A2"] = "Solicitud"
    sheet["B2"] = link.rfq.rfq_number
    sheet["A3"] = "Proveedor"
    sheet["B3"] = link.supplier.name if link.supplier else str(link.supplier_id)
    sheet["A4"] = "Folio de cotizacion *"
    sheet["A5"] = "Vigencia (AAAA-MM-DD)"
    sheet["A6"] = "Moneda"
    sheet["B6"] = "MXN"
    sheet["A7"] = "Dias de entrega"
    sheet["A8"] = "Dias de credito"
    sheet["B8"] = 30
    sheet["A9"] = "Descuento"
    sheet["B9"] = 0
    sheet["A10"] = "Flete"
    sheet["B10"] = 0
    sheet["A11"] = "Impuestos"
    sheet["B11"] = 0

    headers = [
        "ID partida",
        "Codigo",
        "Material",
        "Cantidad",
        "Unidad",
        "Precio unitario *",
        "Dias entrega",
        "Notas",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=ITEM_HEADER_ROW, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="287DB2")
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(link.rfq.items, start=ITEM_HEADER_ROW + 1):
        sheet.cell(row=row, column=1, value=item.id)
        sheet.cell(row=row, column=2, value=item.source_code or "")
        sheet.cell(row=row, column=3, value=item.description)
        sheet.cell(row=row, column=4, value=float(item.quantity))
        sheet.cell(row=row, column=5, value=item.unit)

    widths = {"A": 14, "B": 18, "C": 46, "D": 14, "E": 12, "F": 18, "G": 16, "H": 36}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = f"A{ITEM_HEADER_ROW + 1}"
    sheet.auto_filter.ref = f"A{ITEM_HEADER_ROW}:H{ITEM_HEADER_ROW + len(link.rfq.items)}"
    sheet.protection.sheet = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_quote_template(content: bytes, link: SupplierRFQSupplier) -> SupplierQuoteDraftInput:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("No fue posible leer la plantilla Excel") from exc
    if "_ACSM" not in workbook.sheetnames or "Cotizacion" not in workbook.sheetnames:
        raise ValueError("El archivo no es una plantilla de cotizacion ACSM")

    metadata = workbook["_ACSM"]
    if metadata["A1"].value != TEMPLATE_MARKER or str(metadata["B1"].value) != TEMPLATE_VERSION:
        raise ValueError("La version de la plantilla no es compatible")
    if metadata["B2"].value != link.rfq.rfq_number:
        raise ValueError("La plantilla pertenece a otra solicitud")
    if int(metadata["B3"].value) != link.supplier_id:
        raise ValueError("La plantilla pertenece a otro proveedor")

    sheet = workbook["Cotizacion"]
    quote_number = str(sheet["B4"].value or "").strip()
    if not quote_number:
        raise ValueError("Captura el folio de cotizacion en la plantilla")

    valid_until_value = sheet["B5"].value
    if isinstance(valid_until_value, date):
        valid_until = valid_until_value
    elif valid_until_value:
        try:
            valid_until = date.fromisoformat(str(valid_until_value).strip()[:10])
        except ValueError as exc:
            raise ValueError("La vigencia debe usar el formato AAAA-MM-DD") from exc
    else:
        valid_until = None

    items: list[dict] = []
    for row in sheet.iter_rows(min_row=ITEM_HEADER_ROW + 1, values_only=True):
        rfq_item_id = row[0]
        unit_price = row[5]
        if rfq_item_id in (None, ""):
            continue
        if unit_price in (None, ""):
            continue
        items.append(
            {
                "rfq_item_id": int(rfq_item_id),
                "unit_price": _decimal(unit_price),
                "delivery_days": _integer(row[6]),
                "notes": str(row[7]).strip() if row[7] not in (None, "") else None,
            }
        )

    if not items:
        raise ValueError("La plantilla no contiene precios unitarios")

    return SupplierQuoteDraftInput.model_validate(
        {
            "quote_number": quote_number,
            "valid_until": valid_until,
            "currency": str(sheet["B6"].value or "MXN").strip().upper(),
            "delivery_days": _integer(sheet["B7"].value),
            "payment_terms_days": _integer(sheet["B8"].value, default=30),
            "discount": _decimal(sheet["B9"].value),
            "shipping_cost": _decimal(sheet["B10"].value),
            "tax_amount": _decimal(sheet["B11"].value),
            "items": items,
        }
    )
